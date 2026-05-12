import io
import itertools
import hashlib
import re
from copy import copy
from math import ceil
from collections import Counter, defaultdict
from datetime import datetime, timezone

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import TableColumn
from rapidfuzz import fuzz



SHIP_DEFAULTS = {
    "entity_column": "Name of Vessel",
    "year_column": "Year",
    "type_column": "Type of Veseel",
    "amount_column": "Amount (primary)",
    "unit_column": "Unit (primary)",
    "notes_column_1": "Remarks from ledger",
    "notes_column_2": "Notes from transcriber",
}

EVIDENCE_KIND_OPTIONS = {"categorical", "numeric", "year/date", "text", "auto"}
LEGACY_EVIDENCE_DEFAULTS = {
    "year_column": {"kind": "year/date", "weight": 0.10},
    "type_column": {"kind": "categorical", "weight": 0.08},
    "amount_column": {"kind": "numeric", "weight": 0.05},
    "unit_column": {"kind": "categorical", "weight": 0.12},
    "notes_column_1": {"kind": "text", "weight": 0.03},
    "notes_column_2": {"kind": "text", "weight": 0.03},
}

def normalize_spaces(text: str) -> str:
    return re.sub('\\s+', ' ', text).strip()

def clean_name(value: str) -> str:
    if pd.isna(value):
        return ''
    s = str(value).strip().lower()
    s = s.replace('&', ' and ')
    s = s.replace('?', ' ')
    s = re.sub('[\\.,;:\'\\"`´‘’“”()\\[\\]{}_/\\\\|-]+', ' ', s)
    s = normalize_spaces(s)
    return s

def stable_hash(text: str, length: int=12) -> str:
    return hashlib.sha1(text.encode('utf-8')).hexdigest()[:length]

def clamp_weight(weight, default: float=0.08) -> float:
    try:
        value = float(weight)
    except Exception:
        value = default
    return max(0.0, min(1.0, value))

def infer_evidence_kind_from_series(series: pd.Series, column_name: str='') -> str:
    name = str(column_name or '').strip().lower()
    name_has_year_hint = bool(re.search(r'(year|date|day|month)', name))
    name_has_categorical_hint = bool(re.search(r'(type|category|unit|class|kind)', name))
    name_has_text_hint = bool(re.search(r'(note|remark|comment|description)', name))
    if name_has_year_hint:
        return 'year/date'
    if name_has_categorical_hint:
        return 'categorical'
    if pd.api.types.is_numeric_dtype(series):
        return 'numeric'
    nonempty = series.dropna().astype(str).str.strip()
    if nonempty.empty:
        return 'categorical'
    parsed_numeric = pd.to_numeric(nonempty, errors='coerce')
    numeric_ratio = float(parsed_numeric.notna().mean()) if len(parsed_numeric) else 0.0
    if numeric_ratio >= 0.9 and nonempty.str.contains(r"\d", regex=True).mean() > 0.8:
        return 'numeric'
    if name_has_text_hint:
        return 'text'
    sample = nonempty.head(200)
    avg_len = float(sample.map(len).mean()) if len(sample) else 0.0
    cardinality = float(nonempty.nunique(dropna=True)) / max(1, len(nonempty))
    if avg_len > 40:
        return 'text'
    if cardinality <= 0.2 and nonempty.nunique(dropna=True) <= 50:
        return 'categorical'
    return 'text'

def sanitize_evidence_field(field: dict, idx: int=0, infer_kind: str='categorical') -> dict | None:
    if not isinstance(field, dict):
        return None
    column = str(field.get('column') or '').strip()
    if not column:
        return None
    raw_kind = str(field.get('kind') or 'auto').strip().lower()
    kind = raw_kind if raw_kind in EVIDENCE_KIND_OPTIONS else 'auto'
    stable_id = str(field.get('id') or f"E{stable_hash(f'{column}::{kind}::{idx}', 10)}")
    return {
        'id': stable_id,
        'column': column,
        'kind': kind,
        'resolved_kind': infer_kind if kind == 'auto' else kind,
        'weight': clamp_weight(field.get('weight'), default=0.08),
        'enabled': bool(field.get('enabled', True)),
    }

def ensure_unique_evidence_ids(fields: list[dict], entity_column: str) -> list[dict]:
    seen_ids = set()
    out = []
    for idx, field in enumerate(fields):
        clean_field = dict(field)
        candidate_id = str(clean_field.get('id') or '').strip()
        if not candidate_id or candidate_id in seen_ids:
            unique_seed = f"unique::{entity_column}::{clean_field.get('column', '')}::{idx}::{len(seen_ids)}"
            candidate_id = f"E{stable_hash(unique_seed, 12)}"
            while candidate_id in seen_ids:
                candidate_id = f"E{stable_hash(candidate_id + 'x', 12)}"
        clean_field['id'] = candidate_id
        seen_ids.add(candidate_id)
        out.append(clean_field)
    return out

def normalize_column_config(column_config: dict | None, available_columns: list[str] | None=None) -> dict:
    cfg = dict(column_config or {})
    entity_column = cfg.get('entity_column') or SHIP_DEFAULTS['entity_column']
    evidence_fields = []
    seen = set()
    raw_evidence_fields = cfg.get('evidence_fields')
    if isinstance(raw_evidence_fields, list):
        for idx, field in enumerate(raw_evidence_fields):
            clean_field = sanitize_evidence_field(field, idx)
            if not clean_field:
                continue
            key = (clean_field['column'], clean_field['resolved_kind'])
            if key in seen:
                continue
            if available_columns is not None and clean_field['column'] not in available_columns:
                continue
            seen.add(key)
            evidence_fields.append(clean_field)
    for legacy_key, meta in LEGACY_EVIDENCE_DEFAULTS.items():
        legacy_col = cfg.get(legacy_key)
        if not legacy_col:
            continue
        legacy_col = str(legacy_col).strip()
        if not legacy_col:
            continue
        if available_columns is not None and legacy_col not in available_columns:
            continue
        key = (legacy_col, meta['kind'])
        if key in seen:
            continue
        seen.add(key)
        evidence_fields.append({
            'id': f"E{stable_hash(f'legacy::{legacy_key}::{legacy_col}', 10)}",
            'column': legacy_col,
            'kind': meta['kind'],
            'resolved_kind': meta['kind'],
            'weight': meta['weight'],
            'enabled': True,
        })
    evidence_fields = ensure_unique_evidence_ids(evidence_fields, entity_column=entity_column)
    return {
        'entity_column': entity_column,
        'evidence_fields': evidence_fields,
    }

def suggest_evidence_fields(df: pd.DataFrame, entity_column: str, max_fields: int=6) -> list[dict]:
    candidates = []
    for idx, column in enumerate(df.columns.tolist()):
        if column == entity_column:
            continue
        series = df[column]
        nonempty = series.dropna()
        if nonempty.empty:
            continue
        kind = infer_evidence_kind_from_series(series, column)
        name = str(column).lower()
        coverage = float(nonempty.shape[0]) / max(1, len(series))
        score = coverage
        if kind == 'year/date':
            score += 0.35
        elif kind == 'categorical':
            cardinality = float(nonempty.astype(str).str.strip().nunique()) / max(1, len(nonempty))
            if cardinality <= 0.3:
                score += 0.25
        elif kind == 'numeric':
            score += 0.20
        elif kind == 'text':
            score += 0.05
        if re.search(r'(note|remark|comment)', name):
            score -= 0.1
        candidates.append((score, idx, column, kind))
    candidates.sort(key=lambda x: (-x[0], x[1]))
    defaults = {'year/date': 0.10, 'categorical': 0.08, 'numeric': 0.05, 'text': 0.03}
    out = []
    for rank, (_, _, column, kind) in enumerate(candidates[:max_fields], start=1):
        out.append({
            'id': f"E{stable_hash(f'suggest::{column}::{kind}::{rank}', 10)}",
            'column': column,
            'kind': kind,
            'resolved_kind': kind,
            'weight': defaults.get(kind, 0.08),
            'enabled': True,
        })
    return out

def make_pair_key(name_a: str, name_b: str, entity_column: str='Name of Vessel') -> str:
    normalized_names = sorted([clean_name(name_a), clean_name(name_b)])
    raw_key = entity_column + '::' + '||'.join(normalized_names)
    return 'P' + stable_hash(raw_key)

def make_auto_group_key(strict_key: str, entity_column: str) -> str:
    return 'A' + stable_hash(f'auto::{entity_column}::{strict_key}')

def strict_name_key(value: str) -> str:
    if pd.isna(value):
        return ''
    s = str(value).strip().lower()
    s = re.sub('[^a-z0-9]+', '', s)
    return s

def clean_vessel_type(value: str) -> str:
    if pd.isna(value):
        return ''
    s = str(value).strip().lower()
    s = re.sub('[\\.,;:\'\\"`´‘’“”()\\[\\]{}_/\\\\|-]+', ' ', s)
    s = normalize_spaces(s)
    mapping = {'prop': 'propeller', 'ss': 'steamship', 's s': 'steamship', 'stmr': 'steamer', 'stm': 'steamer', 'str': 'steamer', 'sch': 'schooner', 'schr': 'schooner', 'shnr': 'schooner', 'sb': 'sailboat', 's b': 'sailboat'}
    return mapping.get(s, s)

def clean_unit(value: str) -> str:
    if pd.isna(value):
        return ''
    s = str(value).strip().lower()
    s = re.sub('[\\.,;:\'\\"`´‘’“”()\\[\\]{}_/\\\\|-]+', ' ', s)
    s = normalize_spaces(s)
    mapping = {'ton': 'tons', 'tons': 'tons', 'tns': 'tons', 'toise': 'toise', 'barrels': 'barrels', 'barrel': 'barrels', 'baskets': 'baskets', 'crates': 'crates', 'boxes': 'boxes'}
    return mapping.get(s, s)

def clean_categorical_value(value: str, column_name: str='') -> str:
    name = str(column_name or '').lower()
    if 'unit' in name:
        return clean_unit(value)
    if 'type' in name or 'category' in name:
        return clean_vessel_type(value)
    return clean_name(value)

def to_float(value):
    if pd.isna(value):
        return None
    s = str(value).strip().replace(',', '')
    match = re.search('-?\\d+(?:\\.\\d+)?', s)
    if not match:
        return None
    try:
        return float(match.group())
    except Exception:
        return None

def parse_year_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return int(value.year)
    if isinstance(value, datetime):
        return int(value.year)
    s = str(value).strip()
    if not s:
        return None
    match = re.search(r'\b(1[5-9]\d{2}|20\d{2})\b', s)
    if match:
        return int(match.group(1))
    numeric = pd.to_numeric(pd.Series([s]), errors='coerce').iloc[0]
    if pd.notna(numeric):
        numeric_int = int(numeric)
        if 1500 <= numeric_int <= 2100:
            return numeric_int
    return None

def most_common_nonempty(values, default=''):
    vals = [v for v in values if v]
    if not vals:
        return default
    return Counter(vals).most_common(1)[0][0]

def set_to_text(values, limit=6):
    vals = sorted((v for v in set(values) if v))
    if not vals:
        return ''
    shown = vals[:limit]
    suffix = '' if len(vals) <= limit else f' (+{len(vals) - limit} more)'
    return ' | '.join(shown) + suffix

def year_overlap_score(a_min, a_max, b_min, b_max):
    if pd.isna(a_min) or pd.isna(a_max) or pd.isna(b_min) or pd.isna(b_max):
        return 0.5
    if max(a_min, b_min) <= min(a_max, b_max):
        return 1.0
    gap = min(abs(a_min - b_max), abs(b_min - a_max))
    if gap <= 1:
        return 0.8
    if gap <= 3:
        return 0.5
    if gap <= 8:
        return 0.2
    return 0.0

def overlap_score(a_set, b_set, missing_default=0.5):
    if not a_set and (not b_set):
        return missing_default
    if not a_set or not b_set:
        return missing_default
    return 1.0 if set(a_set) & set(b_set) else 0.0

def tons_similarity(a_value, b_value):
    if a_value is None or b_value is None:
        return 0.5
    max_val = max(abs(a_value), abs(b_value), 1.0)
    rel_diff = abs(a_value - b_value) / max_val
    if rel_diff <= 0.05:
        return 1.0
    if rel_diff <= 0.15:
        return 0.8
    if rel_diff <= 0.3:
        return 0.5
    if rel_diff <= 0.5:
        return 0.2
    return 0.0

def build_reason_list(name_score, same_clean, evidence_details=None):
    reasons = []
    if same_clean:
        reasons.append('same cleaned name')
    if name_score >= 0.95:
        reasons.append('very high raw-name similarity')
    elif name_score >= 0.88:
        reasons.append('high raw-name similarity')
    for detail in evidence_details or []:
        if detail.get('score', 0.0) >= 0.8:
            reasons.append(detail.get('reason', f"{detail.get('column', 'evidence')}: strong match"))
    if not reasons:
        reasons.append('flagged by fuzzy candidate generation')
    return reasons

def canonical_sort_key(name: str):
    whitespace_normalized = normalize_spaces(str(name))
    compact_alnum = re.sub('[^A-Za-z0-9]', '', whitespace_normalized)
    letters_only = re.sub('[^A-Za-z]', '', whitespace_normalized)
    all_caps = int(bool(letters_only) and letters_only.isupper())
    punctuation_count = len(re.findall('[^\\w\\s]', whitespace_normalized))
    cleaned_for_length = re.sub('[^A-Za-z0-9\\s]', '', whitespace_normalized)
    cleaned_for_length = normalize_spaces(cleaned_for_length)
    return (
        all_caps,
        punctuation_count,
        len(cleaned_for_length),
        whitespace_normalized.lower(),
        compact_alnum.lower(),
    )

def choose_canonical_name(stats_df, name_list):
    subset = stats_df[stats_df['raw_name'].isin(name_list)].copy()
    subset = subset.sort_values(['row_count', 'raw_name'], ascending=[False, True])
    if subset.empty:
        return sorted(name_list, key=canonical_sort_key)[0]
    max_rows = subset['row_count'].max()
    candidates = subset[subset['row_count'] == max_rows]['raw_name'].tolist()
    return sorted(candidates, key=canonical_sort_key)[0]

def preprocess_rows(df: pd.DataFrame, column_config: dict) -> pd.DataFrame:
    normalized_config = normalize_column_config(column_config, available_columns=df.columns.tolist())
    out = df.copy()
    entity_col = normalized_config['entity_column']
    out['raw_name'] = out.get(entity_col, pd.Series('', index=out.index)).fillna('').astype(str).map(lambda x: x.strip())
    out = out[out['raw_name'] != ''].copy()
    out['clean_name'] = out['raw_name'].map(clean_name)
    out['strict_name_key'] = out['raw_name'].map(strict_name_key)

    runtime_fields = []
    for idx, field in enumerate(normalized_config.get('evidence_fields', [])):
        if not field.get('enabled', True):
            continue
        column = field['column']
        if column not in out.columns:
            continue
        source_series = out[column]
        resolved_kind = field.get('resolved_kind', field.get('kind', 'categorical'))
        if field.get('kind') == 'auto':
            resolved_kind = infer_evidence_kind_from_series(source_series, column)
        base_key = f"ev_{stable_hash(field['id'], 10)}"
        runtime_field = {
            'id': field['id'],
            'column': column,
            'kind': field.get('kind', resolved_kind),
            'resolved_kind': resolved_kind,
            'weight': clamp_weight(field.get('weight'), default=0.08),
            'enabled': True,
            'base_key': base_key,
        }
        if resolved_kind == 'categorical':
            out[f"{base_key}_cat"] = source_series.map(lambda v: clean_categorical_value(v, column))
        elif resolved_kind == 'numeric':
            out[f"{base_key}_num"] = source_series.map(to_float)
        elif resolved_kind == 'year/date':
            out[f"{base_key}_year"] = source_series.map(parse_year_value)
        else:
            out[f"{base_key}_txt"] = source_series.fillna('').astype(str).map(lambda v: normalize_spaces(v) if v.strip() else '')
        runtime_fields.append(runtime_field)

    # Legacy compatibility fields retained for summary cards and safe-auto evidence display.
    year_field = next((f for f in runtime_fields if f['resolved_kind'] == 'year/date'), None)
    type_field = next((f for f in runtime_fields if f['resolved_kind'] == 'categorical' and re.search(r'(type|category)', f['column'], re.I)), None)
    unit_field = next((f for f in runtime_fields if f['resolved_kind'] == 'categorical' and re.search(r'unit', f['column'], re.I)), None)
    amount_field = next((f for f in runtime_fields if f['resolved_kind'] == 'numeric' and re.search(r'(amount|ton|weight)', f['column'], re.I)), None)
    text_fields = [f for f in runtime_fields if f['resolved_kind'] == 'text']

    out['year_num'] = out[f"{year_field['base_key']}_year"] if year_field else pd.Series(None, index=out.index)
    out['vessel_type_clean'] = out[f"{type_field['base_key']}_cat"] if type_field else pd.Series('', index=out.index)
    out['unit_primary_clean'] = out[f"{unit_field['base_key']}_cat"] if unit_field else pd.Series('', index=out.index)
    out['amount_primary_num'] = out[f"{amount_field['base_key']}_num"] if amount_field else pd.Series(None, index=out.index)
    if text_fields:
        text_cols = [out[f"{field['base_key']}_txt"] for field in text_fields[:2]]
        combined = text_cols[0].fillna('').astype(str).str.strip()
        if len(text_cols) > 1:
            combined = (combined + ' || ' + text_cols[1].fillna('').astype(str).str.strip()).str.strip(' |')
        out['notes_combined'] = combined
    else:
        out['notes_combined'] = pd.Series('', index=out.index)

    out.attrs['column_config_normalized'] = normalized_config
    out.attrs['evidence_fields_runtime'] = runtime_fields
    return out

def build_name_stats(rows: pd.DataFrame) -> pd.DataFrame:
    runtime_fields = rows.attrs.get('evidence_fields_runtime', [])
    records = []
    for raw_name, grp in rows.groupby('raw_name', sort=False):
        years = grp['year_num'].dropna()
        tons_values = grp.loc[grp['unit_primary_clean'] == 'tons', 'amount_primary_num'].dropna().tolist()
        evidence_agg = {}
        for field in runtime_fields:
            base_key = field['base_key']
            kind = field['resolved_kind']
            if kind == 'categorical':
                values = tuple(sorted(v for v in set(grp[f"{base_key}_cat"].dropna().tolist()) if v))
                evidence_agg[field['id']] = {
                    'column': field['column'],
                    'kind': kind,
                    'weight': field['weight'],
                    'values': values,
                }
            elif kind == 'numeric':
                nums = grp[f"{base_key}_num"].dropna().tolist()
                evidence_agg[field['id']] = {
                    'column': field['column'],
                    'kind': kind,
                    'weight': field['weight'],
                    'median': float(pd.Series(nums).median()) if nums else None,
                }
            elif kind == 'year/date':
                vals = grp[f"{base_key}_year"].dropna().tolist()
                evidence_agg[field['id']] = {
                    'column': field['column'],
                    'kind': kind,
                    'weight': field['weight'],
                    'min_year': int(min(vals)) if vals else None,
                    'max_year': int(max(vals)) if vals else None,
                }
            else:
                texts = [t for t in grp[f"{base_key}_txt"].dropna().astype(str).tolist() if t]
                evidence_agg[field['id']] = {
                    'column': field['column'],
                    'kind': kind,
                    'weight': field['weight'],
                    'sample_text': ' | '.join(texts[:3]),
                }
        records.append({'raw_name': raw_name, 'display_name': raw_name, 'clean_name': most_common_nonempty(grp['clean_name'].tolist(), default=clean_name(raw_name)), 'strict_name_key': most_common_nonempty(grp['strict_name_key'].tolist(), default=strict_name_key(raw_name)), 'row_count': int(len(grp)), 'min_year': int(years.min()) if len(years) else None, 'max_year': int(years.max()) if len(years) else None, 'vessel_types': tuple(sorted((v for v in set(grp['vessel_type_clean']) if v))), 'units': tuple(sorted((v for v in set(grp['unit_primary_clean']) if v))), 'median_tons': float(pd.Series(tons_values).median()) if tons_values else None, 'sample_notes': ' | '.join([n for n in grp['notes_combined'].dropna().astype(str).tolist() if n][:3]), 'evidence_agg': evidence_agg})
    stats = pd.DataFrame(records)
    stats['prefix_block'] = stats['clean_name'].map(lambda s: s[:4] if s else '')
    stats.attrs['evidence_fields_runtime'] = runtime_fields
    return stats

def build_safe_auto_groups(stats: pd.DataFrame, entity_column: str) -> pd.DataFrame:
    groups = []
    for strict_key, grp in stats.groupby('strict_name_key', sort=False):
        if not strict_key or len(grp) < 2:
            continue
        names = grp['raw_name'].tolist()
        canonical = choose_canonical_name(stats, names)
        reasons = ['same strict key after removing punctuation/case/spaces']
        if len(set(grp['clean_name'])) == 1:
            reasons.append('same cleaned name')
        years_min = grp['min_year'].dropna()
        years_max = grp['max_year'].dropna()
        groups.append({'auto_group_id': f'A{len(groups) + 1:04d}', 'auto_group_key': make_auto_group_key(strict_key, entity_column), 'strict_name_key': strict_key, 'canonical_name': canonical, 'member_count': len(grp), 'member_names': ' | '.join(sorted(names)), 'members_list': sorted(names), 'reasons': ' | '.join(reasons), 'confidence': 'safe_auto', 'total_rows': int(grp['row_count'].sum()), 'min_year': int(years_min.min()) if len(years_min) else None, 'max_year': int(years_max.max()) if len(years_max) else None, 'units': set_to_text(itertools.chain.from_iterable(grp['units'].tolist())), 'vessel_types': set_to_text(itertools.chain.from_iterable(grp['vessel_types'].tolist()))})
    out = pd.DataFrame(groups)
    if not out.empty:
        out = out.sort_values(['member_count', 'total_rows', 'canonical_name'], ascending=[False, False, True]).reset_index(drop=True)
    return out

def generate_candidate_pairs(stats: pd.DataFrame, entity_column: str, resolved_names=None, fuzzy_threshold: int=88, column_config: dict | None=None) -> pd.DataFrame:
    if resolved_names is None:
        resolved_names = set()
    runtime_fields = stats.attrs.get('evidence_fields_runtime', [])
    runtime_by_id = {field.get('id'): field for field in runtime_fields if field.get('id')}
    runtime_by_column = {}
    for field in runtime_fields:
        col = field.get('column')
        if col and col not in runtime_by_column:
            runtime_by_column[col] = field
    normalized_config = normalize_column_config(column_config or {"entity_column": entity_column})
    active_evidence_fields = [f for f in normalized_config.get('evidence_fields', []) if f.get('enabled', True) and f.get('weight', 0.0) > 0]
    if column_config is None and not active_evidence_fields:
        active_evidence_fields = [
            {
                'id': field['id'],
                'column': field['column'],
                'kind': field.get('kind', field.get('resolved_kind', 'categorical')),
                'resolved_kind': field.get('resolved_kind', 'categorical'),
                'weight': field.get('weight', 0.08),
                'enabled': True,
            }
            for field in runtime_fields
            if field.get('enabled', True) and field.get('weight', 0.0) > 0
        ]
    resolved_active_fields = []
    for field in active_evidence_fields:
        clean_field = dict(field)
        runtime_field = runtime_by_id.get(clean_field.get('id')) or runtime_by_column.get(clean_field.get('column'))
        if runtime_field:
            clean_field['resolved_kind'] = runtime_field.get('resolved_kind', clean_field.get('resolved_kind', clean_field.get('kind', 'categorical')))
        elif clean_field.get('kind') == 'auto':
            clean_field['resolved_kind'] = 'categorical'
        clean_field['weight'] = clamp_weight(clean_field.get('weight'), default=0.08)
        resolved_active_fields.append(clean_field)
    active_evidence_fields = resolved_active_fields
    candidate_stats = stats[~stats['raw_name'].isin(resolved_names)].copy()
    by_name = {row['raw_name']: row for _, row in candidate_stats.iterrows()}
    pairs = {}
    candidate_records = []
    threshold_score = fuzzy_threshold / 100.0
    lower_name_floor = max(0.78, threshold_score - 0.08)

    def evaluate_evidence_scores(row_a, row_b):
        evidence_scores = []
        for field in active_evidence_fields:
            field_id = field['id']
            field_kind = field.get('resolved_kind', field.get('kind', 'categorical'))
            agg_a = row_a.get('evidence_agg', {}).get(field_id, {})
            agg_b = row_b.get('evidence_agg', {}).get(field_id, {})
            score = 0.5
            reason = f"{field['column']}: neutral evidence"
            if field_kind == 'categorical':
                score = overlap_score(tuple(agg_a.get('values', ())), tuple(agg_b.get('values', ())), missing_default=0.5)
                if score == 1.0:
                    reason = f"{field['column']}: same category"
                elif score == 0.0:
                    reason = f"{field['column']}: conflicting categories"
            elif field_kind == 'numeric':
                score = tons_similarity(agg_a.get('median'), agg_b.get('median'))
                if score >= 0.8:
                    reason = f"{field['column']}: similar median numeric value"
                elif score == 0.0:
                    reason = f"{field['column']}: numeric mismatch"
            elif field_kind == 'year/date':
                score = year_overlap_score(agg_a.get('min_year'), agg_a.get('max_year'), agg_b.get('min_year'), agg_b.get('max_year'))
                if score >= 0.8:
                    reason = f"{field['column']}: overlapping/near years"
                elif score == 0.0:
                    reason = f"{field['column']}: distant years"
            elif field_kind == 'text':
                text_a = agg_a.get('sample_text', '')
                text_b = agg_b.get('sample_text', '')
                if text_a and text_b:
                    score = fuzz.WRatio(text_a, text_b) / 100.0
                else:
                    score = 0.5
                if score >= 0.8:
                    reason = f"{field['column']}: similar text context"
                elif score <= 0.2 and text_a and text_b:
                    reason = f"{field['column']}: different text context"
            evidence_scores.append({
                'id': field_id,
                'column': field['column'],
                'kind': field_kind,
                'weight': clamp_weight(field.get('weight'), default=0.08),
                'score': round(float(score), 4),
                'reason': reason,
            })
        if evidence_scores:
            evidence_weight_total = sum(item['weight'] for item in evidence_scores) or 1.0
            evidence_signal = sum(item['score'] * item['weight'] for item in evidence_scores) / evidence_weight_total
        else:
            evidence_signal = 0.5
        return (evidence_scores, evidence_signal)

    def add_pair(name_a, name_b, source):
        if name_a == name_b:
            return
        if name_a not in by_name or name_b not in by_name:
            return
        key = tuple(sorted((name_a, name_b)))
        if key in pairs:
            pairs[key].add(source)
            return
        pairs[key] = {source}
    for _, grp in candidate_stats.groupby('clean_name'):
        names = grp['raw_name'].tolist()
        if len(names) < 2:
            continue
        for a, b in itertools.combinations(names, 2):
            add_pair(a, b, 'same_clean_name')
    for _, grp in candidate_stats.groupby('prefix_block'):
        names = grp['raw_name'].tolist()
        if len(names) < 2:
            continue
        if len(names) > 60:
            grp = grp.sort_values(['row_count', 'raw_name'], ascending=[False, True]).head(60)
            names = grp['raw_name'].tolist()
        for a, b in itertools.combinations(names, 2):
            row_a = by_name[a]
            row_b = by_name[b]
            raw_score = fuzz.WRatio(row_a['raw_name'], row_b['raw_name']) / 100.0
            clean_score = fuzz.WRatio(row_a['clean_name'], row_b['clean_name']) / 100.0
            if max(raw_score, clean_score) >= threshold_score:
                add_pair(a, b, 'fuzzy_prefix_block')

    # Evidence-assisted expansion for borderline name similarity.
    if active_evidence_fields and len(candidate_stats) > 1:
        blocking_fields = [
            field for field in sorted(active_evidence_fields, key=lambda f: f.get('weight', 0.0), reverse=True)
            if field.get('resolved_kind', field.get('kind')) in {'categorical', 'numeric', 'year/date'}
        ][:3]
        blocks = defaultdict(set)
        max_block_size = 36
        max_candidates_per_block = 28
        for _, row in candidate_stats.iterrows():
            name = row['raw_name']
            evidence_agg = row.get('evidence_agg', {})
            for field in blocking_fields:
                field_id = field['id']
                field_kind = field.get('resolved_kind', field.get('kind'))
                agg = evidence_agg.get(field_id, {})
                if field_kind == 'categorical':
                    for value in tuple(agg.get('values', ())):
                        if not value:
                            continue
                        blocks[(field_id, 'cat', value)].add(name)
                elif field_kind == 'year/date':
                    min_year = agg.get('min_year')
                    max_year = agg.get('max_year')
                    if min_year is None or max_year is None:
                        continue
                    center = int((min_year + max_year) / 2)
                    bucket = center // 2
                    blocks[(field_id, 'year', bucket)].add(name)
                elif field_kind == 'numeric':
                    median = agg.get('median')
                    if median is None:
                        continue
                    bucket = int(round(float(median) / 10.0))
                    blocks[(field_id, 'num', bucket)].add(name)
        for _, names_set in blocks.items():
            if len(names_set) < 2:
                continue
            if len(names_set) > max_block_size:
                continue
            names = sorted(
                names_set,
                key=lambda n: (
                    -int(by_name[n].get('row_count', 0)),
                    n.lower(),
                ),
            )[:max_candidates_per_block]
            for a, b in itertools.combinations(names, 2):
                if (a, b) in pairs or (b, a) in pairs:
                    continue
                row_a = by_name[a]
                row_b = by_name[b]
                raw_score = fuzz.WRatio(row_a['raw_name'], row_b['raw_name']) / 100.0
                clean_score = fuzz.WRatio(row_a['clean_name'], row_b['clean_name']) / 100.0
                name_gate = max(raw_score, clean_score)
                if name_gate >= threshold_score:
                    continue
                if name_gate < lower_name_floor:
                    continue
                evidence_scores, evidence_signal = evaluate_evidence_scores(row_a, row_b)
                strong_evidence_count = sum(1 for item in evidence_scores if item['score'] >= 0.8)
                if evidence_signal >= 0.80 and strong_evidence_count >= 2:
                    add_pair(a, b, 'evidence_assisted_borderline')
    for (a, b), sources in pairs.items():
        row_a = by_name[a]
        row_b = by_name[b]
        raw_score = fuzz.WRatio(row_a['raw_name'], row_b['raw_name']) / 100.0
        clean_score = fuzz.WRatio(row_a['clean_name'], row_b['clean_name']) / 100.0
        same_clean = row_a['clean_name'] == row_b['clean_name']
        name_signal = 0.75 * max(raw_score, clean_score) + 0.25 * (1.0 if same_clean else clean_score)
        evidence_scores, evidence_signal = evaluate_evidence_scores(row_a, row_b)
        if evidence_scores:
            evidence_weight_total = sum(item['weight'] for item in evidence_scores) or 1.0
            evidence_signal = sum(item['score'] * item['weight'] for item in evidence_scores) / evidence_weight_total
            final_score = 0.65 * name_signal + 0.35 * evidence_signal
        else:
            evidence_signal = 0.5
            final_score = name_signal
        year_score = year_overlap_score(row_a['min_year'], row_a['max_year'], row_b['min_year'], row_b['max_year'])
        unit_score = overlap_score(row_a['units'], row_b['units'])
        type_score = overlap_score(row_a['vessel_types'], row_b['vessel_types'])
        cargo_amount_score = tons_similarity(row_a['median_tons'], row_b['median_tons'])
        pair_key = make_pair_key(row_a['display_name'], row_b['display_name'], entity_column)
        reasons_list = build_reason_list(max(raw_score, clean_score), same_clean, evidence_scores)
        if 'evidence_assisted_borderline' in sources and max(raw_score, clean_score) < threshold_score:
            reasons_list.append('supporting evidence boosted borderline name match')
        candidate_records.append({'pair_key': pair_key, 'candidate_id': pair_key, 'name_a': row_a['display_name'], 'name_b': row_b['display_name'], 'score': round(final_score, 4), 'raw_name_score': round(raw_score, 4), 'clean_name_score': round(clean_score, 4), 'name_signal_score': round(name_signal, 4), 'evidence_signal_score': round(evidence_signal, 4), 'year_score': round(year_score, 4), 'unit_score': round(unit_score, 4), 'type_score': round(type_score, 4), 'cargo_amount_score': round(cargo_amount_score, 4), 'evidence_scores': evidence_scores, 'reasons': ' | '.join(dict.fromkeys(reasons_list)), 'suggested_canonical': choose_canonical_name(candidate_stats, [row_a['display_name'], row_b['display_name']])})
    queue = pd.DataFrame(candidate_records)
    if not queue.empty:
        queue = queue.sort_values(['score'], ascending=[False]).reset_index(drop=True)
    return queue

class UnionFind:

    def __init__(self):
        self.parent = {}

    def find(self, x):
        if x not in self.parent:
            self.parent[x] = x
            return x
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]

    def union(self, a, b):
        ra, rb = (self.find(a), self.find(b))
        if ra != rb:
            self.parent[rb] = ra

def build_merge_outputs(stats_df, auto_groups_df, auto_status, manual_decisions):
    uf = UnionFind()
    history_records = []
    if not auto_groups_df.empty:
        for _, row in auto_groups_df.iterrows():
            gid = row['auto_group_key']
            if auto_status.get(gid) == 'accepted':
                members = row['members_list']
                for m in members[1:]:
                    uf.union(members[0], m)
                history_records.append({'merge_source': 'auto_group', 'merge_id': gid, 'canonical_name': row['canonical_name'], 'members': ' | '.join(members), 'reason': row['reasons'], 'reviewer_comment': '', 'status': 'active'})
    if manual_decisions:
        for pair_key, row in manual_decisions.items():
            if row.get('decision') == 'merge':
                uf.union(row['name_a'], row['name_b'])
                history_records.append({'merge_source': 'manual_pair', 'merge_id': pair_key, 'canonical_name': row['suggested_canonical'], 'members': f"{row['name_a']} | {row['name_b']}", 'reason': row['reasons'], 'reviewer_comment': row.get('reviewer_comment', ''), 'status': 'active'})
    all_names = set()
    for rec in history_records:
        parts = [p.strip() for p in rec['members'].split('|')]
        all_names.update([p for p in parts if p])
    groups = defaultdict(list)
    for name in all_names:
        groups[uf.find(name)].append(name)
    row_count_map = dict(zip(stats_df['raw_name'], stats_df['row_count']))
    clean_map = dict(zip(stats_df['raw_name'], stats_df['clean_name']))
    mapping_records = []
    for idx, members in enumerate(sorted(groups.values(), key=lambda g: (-len(g), sorted(g)[0])), start=1):
        canonical = sorted(members, key=lambda n: (-row_count_map.get(n, 0), *canonical_sort_key(n)))[0]
        cluster_id = f'M{idx:04d}'
        for member in sorted(members):
            mapping_records.append({'cluster_id': cluster_id, 'canonical_name': canonical, 'member_name': member, 'row_count': row_count_map.get(member, 0), 'clean_name': clean_map.get(member, '')})
    return (pd.DataFrame(history_records), pd.DataFrame(mapping_records))

def resolved_names_from_auto(auto_groups_df, auto_status):
    resolved = set()
    if auto_groups_df.empty:
        return resolved
    for _, row in auto_groups_df.iterrows():
        if auto_status.get(row['auto_group_key']) == 'accepted':
            resolved.update(row['members_list'])
    return resolved

def active_manual_decisions_for_config(manual_decisions: dict, column_config: dict) -> dict:
    active_entity = column_config['entity_column']
    return {key: record for key, record in manual_decisions.items() if record.get('entity_column') == active_entity}

def make_download_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode('utf-8')

def build_cleaned_workbook_bytes(raw_df: pd.DataFrame, mapping_df: pd.DataFrame, history_df: pd.DataFrame, manual_decisions: dict, auto_groups_df: pd.DataFrame, column_config: dict, sheet_name: str, workbook_bytes: bytes | None=None, candidate_queue_df: pd.DataFrame | None=None, overwrite_entity_column: bool=False) -> bytes:
    entity_col = column_config['entity_column']
    cleaned_df = raw_df.copy()
    original_values = cleaned_df[entity_col].fillna('').astype(str)
    normalized_original_values = original_values.map(lambda x: str(x).strip())
    canonical_map = {}
    cluster_map = {}
    merged_values = set()
    if not mapping_df.empty:
        canonical_map = dict(zip(mapping_df['member_name'], mapping_df['canonical_name']))
        cluster_map = dict(zip(mapping_df['member_name'], mapping_df['cluster_id']))
        merged_values = set(mapping_df['member_name'].tolist())
    status_map = {}
    source_map = {}
    score_map = {}
    reason_map = {}
    reviewer_comment_map = {}
    if not auto_groups_df.empty:
        accepted = set(history_df.loc[history_df['merge_source'] == 'auto_group', 'merge_id'].tolist()) if not history_df.empty else set()
        for _, row in auto_groups_df.iterrows():
            if row['auto_group_key'] in accepted:
                for member in row['members_list']:
                    status_map[member] = 'merged'
                    source_map[member] = 'auto_group'
                    reason_map[member] = row.get('reasons', '')
                    reviewer_comment_map[member] = row.get('reviewer_comment', '')
    for rec in manual_decisions.values():
        names = [rec.get('name_a', ''), rec.get('name_b', '')]
        decision = rec.get('decision', '')
        reviewer_comment = rec.get('reviewer_comment', '')
        for name in names:
            if name in merged_values or decision == 'merge':
                status_map[name] = 'merged'
                source_map[name] = 'manual_pair'
            elif decision == 'keep_separate':
                status_map[name] = 'reviewed_keep_separate'
                source_map[name] = 'manual_keep_separate'
            elif decision == 'unsure':
                status_map[name] = 'reviewed_unsure'
                source_map[name] = 'manual_unsure'
            score_map[name] = rec.get('score', '')
            reason_map[name] = rec.get('reasons', '')
            reviewer_comment_map[name] = reviewer_comment
    unreviewed_names = set()
    if candidate_queue_df is not None and (not candidate_queue_df.empty):
        for _, row in candidate_queue_df.iterrows():
            unreviewed_names.add(row['name_a'])
            unreviewed_names.add(row['name_b'])
    cleaned_df['dedupe_entity_column'] = entity_col
    cleaned_df['dedupe_original_value'] = original_values
    cleaned_df['dedupe_canonical_value'] = normalized_original_values.map(lambda x: canonical_map.get(x, x))
    cleaned_df['dedupe_cluster_id'] = normalized_original_values.map(lambda x: cluster_map.get(x, ''))
    cleaned_df['dedupe_review_status'] = normalized_original_values.map(lambda x: status_map.get(x, '')).fillna('')
    cleaned_df['dedupe_decision_source'] = normalized_original_values.map(lambda x: source_map.get(x, '')).fillna('')
    cleaned_df['dedupe_score'] = normalized_original_values.map(lambda x: score_map.get(x, '')).fillna('')
    cleaned_df['dedupe_reason'] = normalized_original_values.map(lambda x: reason_map.get(x, '')).fillna('')
    cleaned_df['dedupe_reviewer_comment'] = normalized_original_values.map(lambda x: reviewer_comment_map.get(x, '')).fillna('')

    def default_status(v: str) -> str:
        if v in merged_values:
            return 'merged'
        if v in unreviewed_names:
            return 'unreviewed'
        return 'not_flagged'
    missing_status = cleaned_df['dedupe_review_status'].isna() | (cleaned_df['dedupe_review_status'] == '')
    cleaned_df.loc[missing_status, 'dedupe_review_status'] = normalized_original_values[missing_status].map(default_status)

    dedupe_columns = [
        'dedupe_entity_column',
        'dedupe_original_value',
        'dedupe_canonical_value',
        'dedupe_cluster_id',
        'dedupe_review_status',
        'dedupe_decision_source',
        'dedupe_score',
        'dedupe_reason',
        'dedupe_reviewer_comment',
    ]

    if overwrite_entity_column:
        cleaned_df[entity_col] = cleaned_df['dedupe_canonical_value']

    def to_excel_value(value):
        if pd.isna(value):
            return None
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value

    if workbook_bytes:
        workbook = load_workbook(io.BytesIO(workbook_bytes))
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f'Sheet not found: {sheet_name}')
        worksheet = workbook[sheet_name]

        def normalize_header(value):
            return value.strip() if isinstance(value, str) else value

        header_map = {}
        for col_idx in range(1, worksheet.max_column + 1):
            header_value = worksheet.cell(row=1, column=col_idx).value
            key = normalize_header(header_value)
            if key is not None:
                header_map[key] = col_idx

        existing_dedupe_indexes = [header_map.get(col) for col in dedupe_columns if header_map.get(col) is not None]
        if existing_dedupe_indexes:
            template_col_idx = max(1, min(existing_dedupe_indexes) - 1)
        else:
            template_col_idx = max(1, worksheet.max_column)
        template_letter = get_column_letter(template_col_idx)
        template_dim = worksheet.column_dimensions.get(template_letter)

        dedupe_column_indexes = {}
        for dedupe_col in dedupe_columns:
            existing_idx = header_map.get(dedupe_col)
            if existing_idx is not None:
                dedupe_column_indexes[dedupe_col] = existing_idx
            else:
                new_idx = worksheet.max_column + 1
                dedupe_column_indexes[dedupe_col] = new_idx
                worksheet.cell(row=1, column=new_idx, value=dedupe_col)
                if template_dim is not None:
                    new_letter = get_column_letter(new_idx)
                    worksheet.column_dimensions[new_letter].width = template_dim.width

        dedupe_max_col_idx = max(dedupe_column_indexes.values()) if dedupe_column_indexes else worksheet.max_column

        # If the worksheet uses an Excel table, extend the right edge to include dedupe columns
        # so table styling (header + banded rows) naturally applies to the new columns.
        for table in worksheet.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if min_row != 1:
                continue
            if max_col < template_col_idx:
                continue
            if dedupe_max_col_idx <= max_col:
                continue
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(dedupe_max_col_idx)}{max_row}"
            table.ref = new_ref
            if table.autoFilter is not None:
                table.autoFilter.ref = new_ref

            total_cols = dedupe_max_col_idx - min_col + 1
            existing_cols = list(table.tableColumns)
            updated_cols = []
            for position in range(total_cols):
                col_idx = min_col + position
                header_value = worksheet.cell(row=min_row, column=col_idx).value
                if header_value is None or str(header_value).strip() == '':
                    column_name = f'Column{position + 1}'
                else:
                    column_name = str(header_value)
                if position < len(existing_cols):
                    col_obj = existing_cols[position]
                    col_obj.id = position + 1
                    col_obj.name = column_name
                else:
                    col_obj = TableColumn(id=position + 1, name=column_name)
                updated_cols.append(col_obj)
            table.tableColumns = updated_cols

        data_row_count = len(cleaned_df)
        for dedupe_col, col_idx in dedupe_column_indexes.items():
            template_header = worksheet.cell(row=1, column=template_col_idx)
            target_header = worksheet.cell(row=1, column=col_idx)
            if template_header.has_style:
                target_header._style = copy(template_header._style)

            for row_idx in range(data_row_count):
                excel_row = row_idx + 2
                template_cell = worksheet.cell(row=excel_row, column=template_col_idx)
                target_cell = worksheet.cell(row=excel_row, column=col_idx)
                if template_cell.has_style:
                    target_cell._style = copy(template_cell._style)

            values = cleaned_df[dedupe_col].tolist()
            for row_idx, value in enumerate(values):
                excel_row = row_idx + 2
                worksheet.cell(row=excel_row, column=col_idx, value=to_excel_value(value))

            clear_start = data_row_count + 2
            if clear_start <= worksheet.max_row:
                for excel_row in range(clear_start, worksheet.max_row + 1):
                    worksheet.cell(row=excel_row, column=col_idx, value=None)

        if overwrite_entity_column:
            raw_columns = raw_df.columns.tolist()
            if entity_col in raw_columns:
                entity_col_idx = raw_columns.index(entity_col) + 1
                values = cleaned_df['dedupe_canonical_value'].tolist()
                for row_idx, value in enumerate(values):
                    excel_row = row_idx + 2
                    worksheet.cell(row=excel_row, column=entity_col_idx, value=to_excel_value(value))

        def autofit_columns_and_rows(ws):
            min_width = 8.0
            max_width = 80.0
            col_width_map = {}
            max_row = ws.max_row
            max_col = ws.max_column

            for col_idx in range(1, max_col + 1):
                max_len = 0
                for row_idx in range(1, max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is None:
                        continue
                    text = str(cell_value)
                    longest_segment = max((len(segment) for segment in text.splitlines()), default=0)
                    if longest_segment > max_len:
                        max_len = longest_segment
                width = min(max_width, max(min_width, float(max_len + 2)))
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                col_width_map[col_idx] = width

            base_height = ws.sheet_format.defaultRowHeight or 15.0
            max_height = 180.0
            for row_idx in range(1, max_row + 1):
                max_lines = 1
                for col_idx in range(1, max_col + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is None:
                        continue
                    text = str(cell_value)
                    segments = text.splitlines() if text else ['']
                    usable_width = max(1.0, col_width_map.get(col_idx, min_width) - 2.0)
                    line_count = 0
                    for segment in segments:
                        segment_len = len(segment)
                        line_count += max(1, int(ceil(segment_len / usable_width)))
                    if line_count > max_lines:
                        max_lines = line_count
                ws.row_dimensions[row_idx].height = min(max_height, float(base_height * max_lines))

        autofit_columns_and_rows(worksheet)

        out = io.BytesIO()
        workbook.save(out)
        return out.getvalue()

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        cleaned_df.to_excel(writer, sheet_name=sheet_name, index=False)
    return out.getvalue()

def build_standardized_workbook_bytes(raw_df: pd.DataFrame, mapping_df: pd.DataFrame, column_config: dict, sheet_name: str, workbook_bytes: bytes | None) -> bytes:
    entity_col = column_config['entity_column']
    canonical_map = {}
    if not mapping_df.empty:
        canonical_map = dict(zip(mapping_df['member_name'], mapping_df['canonical_name']))

    standardized_df = raw_df.copy()
    standardized_original_values = standardized_df[entity_col].fillna('').astype(str).map(lambda x: str(x).strip())
    standardized_df[entity_col] = standardized_original_values.map(lambda x: canonical_map.get(x, x))

    def to_excel_value(value):
        if pd.isna(value):
            return None
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        return value

    if workbook_bytes:
        workbook = load_workbook(io.BytesIO(workbook_bytes))
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f'Sheet not found: {sheet_name}')
        worksheet = workbook[sheet_name]

        raw_columns = raw_df.columns.tolist()
        if entity_col not in raw_columns:
            raise KeyError(f'Entity column not found: {entity_col}')
        entity_col_idx = raw_columns.index(entity_col) + 1

        values = standardized_df[entity_col].tolist()
        for row_idx, value in enumerate(values):
            excel_row = row_idx + 2
            worksheet.cell(row=excel_row, column=entity_col_idx, value=to_excel_value(value))

        out = io.BytesIO()
        workbook.save(out)
        return out.getvalue()

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        standardized_df.to_excel(writer, sheet_name=sheet_name, index=False)
    return out.getvalue()

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def build_manual_decision_record(row: pd.Series, decision: str, entity_column: str, reviewer_comment: str=''):
    return {
        'pair_key': row['pair_key'],
        'entity_column': entity_column,
        'name_a': row['name_a'],
        'name_b': row['name_b'],
        'decision': decision,
        'score': float(row['score']),
        'raw_name_score': float(row['raw_name_score']),
        'clean_name_score': float(row['clean_name_score']),
        'name_signal_score': float(row.get('name_signal_score', row['score'])),
        'evidence_signal_score': float(row.get('evidence_signal_score', 0.5)),
        'year_score': float(row.get('year_score', 0.5)),
        'unit_score': float(row.get('unit_score', 0.5)),
        'type_score': float(row.get('type_score', 0.5)),
        'cargo_amount_score': float(row.get('cargo_amount_score', 0.5)),
        'evidence_scores': row.get('evidence_scores', []),
        'reviewer_comment': str(reviewer_comment or ''),
        'reasons': row['reasons'],
        'suggested_canonical': row['suggested_canonical'],
        'updated_at': now_iso(),
    }
